import streamlit as st
import pandas as pd
import json
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime

TEMPLATE_PATH = Path("internal_template.xlsx")
DRIVE_FILENAME = "debtor_data.json"
SCOPES = ["https://www.googleapis.com/auth/drive"]

def get_drive():
    try:
        creds_dict = st.secrets["gdrive"]
        gauth = GoogleAuth()
        gauth.credentials = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scopes=SCOPES)
        return GoogleDrive(gauth)
    except Exception as e:
        st.error(f"Google Drive認証エラー: {e}")
        return None

drive = get_drive()

def load_data_from_drive():
    try:
        file_list = drive.ListFile({'q': f"title='{DRIVE_FILENAME}'"}).GetList()
        if file_list:
            file = file_list[0]
            file.GetContentFile("temp_debtor_data.json")
            with open("temp_debtor_data.json", "r", encoding="utf-8") as f:
                data = json.load(f)
            st.info("Google Driveからデータを読み込みました")
            return data
        else:
            st.warning("Google Driveにデータファイルがありません。")
            return []
    except Exception as e:
        st.error(f"読み込みエラー: {e}")
        return []

def save_data_to_drive(data):
    try:
        with open("temp_debtor_data.json", "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        file_list = drive.ListFile({'q': f"title='{DRIVE_FILENAME}'"}).GetList()
        if file_list:
            file = file_list[0]
            file.SetContentFile("temp_debtor_data.json")
            file.Upload()
            st.info("Google Driveにデータを上書き保存しました")
        else:
            file = drive.CreateFile({'title': DRIVE_FILENAME})
            file.SetContentFile("temp_debtor_data.json")
            file.Upload()
            st.info("Google Driveに新規ファイルを作成しました")
    except Exception as e:
        st.error(f"保存エラー: {e}")

if "uploaded_data" not in st.session_state:
    st.session_state.uploaded_data = load_data_from_drive()
if "json_clear_flag" not in st.session_state:
    st.session_state["json_clear_flag"] = False

st.subheader("債権者データ（JSON）を貼り付けて登録")
if st.session_state["json_clear_flag"]:
    default_json = ""
    st.session_state["json_clear_flag"] = False
else:
    default_json = st.session_state.get("json_input_area", "")

json_input = st.text_area(
    "下にJSONデータを貼り付けてください（配列または単体）",
    value=default_json,
    height=300,
    key="json_input_area"
)
if st.button("JSONを登録"):
    try:
        data = json.loads(st.session_state["json_input_area"])
        if isinstance(data, dict): data = [data]
        st.session_state.uploaded_data.extend(data)
        save_data_to_drive(st.session_state.uploaded_data)
        st.success(f"{len(data)} 件のデータを追加しました。")
        st.session_state["json_clear_flag"] = True
        st.rerun()
    except json.JSONDecodeError as e:
        st.error(f"JSONの形式に誤りがあります: {e}")

if st.session_state.uploaded_data:
    df_all = pd.DataFrame(st.session_state.uploaded_data)
    debtor_names = df_all["debtor_name"].dropna().unique().tolist()
    selected_debtor = st.selectbox("債務者を選択", debtor_names)
    df_debtor = df_all[df_all["debtor_name"] == selected_debtor].reset_index(drop=True)

    st.dataframe(df_debtor, use_container_width=True)

    if not df_debtor.empty:
        st.subheader("行ごと削除")
        for i, row in df_debtor.iterrows():
            cols = st.columns([6, 1])
            with cols[0]:
                st.markdown(f"<b>{row['company_name']}</b>", unsafe_allow_html=True)
            with cols[1]:
                if st.button("削除", key=f"delete_{i}_{selected_debtor}"):
                    mask = (df_all["debtor_name"] == row["debtor_name"])
                    for col in df_debtor.columns:
                        mask = mask & (df_all[col] == row[col])
                    idx_to_drop = df_all[mask].index
                    if len(idx_to_drop) > 0:
                        df_all_new = df_all.drop(idx_to_drop[0]).reset_index(drop=True)
                        st.session_state.uploaded_data = df_all_new.to_dict(orient="records")
                        save_data_to_drive(st.session_state.uploaded_data)
                        st.rerun()
                    else:
                        st.warning("データの一致行が見つかりませんでした。")

        # ▼▼▼ 債務者まるごと削除：確認ダイアログ付き ▼▼▼
        with st.expander(f"{selected_debtor} の全データを削除（注意）"):
            confirm = st.checkbox(f"本当に {selected_debtor} の全データを削除しますか？", key="confirm_delete_debtor")
            if st.button(f"{selected_debtor} の全データを削除", key="delete_debtor_all", disabled=not confirm):
                df_all_new = df_all[df_all["debtor_name"] != selected_debtor].reset_index(drop=True)
                st.session_state.uploaded_data = df_all_new.to_dict(orient="records")
                save_data_to_drive(st.session_state.uploaded_data)
                st.success(f"{selected_debtor} の全データを削除しました。")
                st.rerun()
    else:
        st.info("この債務者のデータはありません。")

    # ここからoutput/YYYYMMDD/に出力
    def make_excel(debtor, df):
        if df.empty:
            st.warning("書き込むデータがありません")
            return None
        df_no_debtor = df.drop(columns=["debtor_name"], errors="ignore")
        today_str = datetime.now().strftime("%Y%m%d")
        output_dir = Path("output") / today_str
        output_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"{today_str}_{debtor}_債権調査票一覧.xlsx"
        output_path = output_dir / file_name
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        ws.cell(row=1, column=3, value=debtor)
        start_row = 4
        start_col = 2
        for row_idx, row in enumerate(df_no_debtor.itertuples(index=False), start=start_row):
            for col_idx, value in enumerate(row, start=start_col):
                ws.cell(row=row_idx, column=col_idx, value=value)
        wb.save(output_path)
        return BytesIO(open(output_path, "rb").read())

    today_str = datetime.now().strftime("%Y%m%d")
    excel = make_excel(selected_debtor, df_debtor)
    if excel:
        st.download_button(
            "Excelダウンロード",
            data=excel,
            file_name=f"{today_str}_{selected_debtor}_債権調査票一覧.xlsx"
        )
